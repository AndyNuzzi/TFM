from pathlib import Path
from typing import Iterable

from openpyxl import Workbook


class ExcelExporter:
    def export(self, datasets: Iterable, output_path: Path) -> Path:
        wb = Workbook()

        ws_docs = wb.active
        ws_docs.title = "documents"

        ws_metrics = wb.create_sheet("entry_profile_metrics")
        ws_subjects = wb.create_sheet("subject_results")

        self._write_documents_sheet(ws_docs, datasets)
        self._write_metrics_sheet(ws_metrics, datasets)
        self._write_subjects_sheet(ws_subjects, datasets)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def _write_documents_sheet(self, ws, datasets):
        headers = [
            "document_id",
            "degree_name",
            "academic_year",
            "family_id",
            "document_type",
            "source_pdf",
        ]
        ws.append(headers)

        for dataset in datasets:
            doc = dataset.document
            ws.append([
                doc.document_id,
                doc.degree_name,
                doc.academic_year,
                doc.family_id.value if hasattr(doc.family_id, "value") else doc.family_id,
                doc.document_type.value if hasattr(doc.document_type, "value") else doc.document_type,
                doc.source_pdf,
            ])

    def _write_metrics_sheet(self, ws, datasets):
        headers = [
            "document_id",
            "degree_name",
            "academic_year",
            "campus",
            "metric_name_std",
            "metric_name_raw",
            "metric_value",
            "unit",
            "source_pdf",
        ]
        ws.append(headers)

        for dataset in datasets:
            for row in dataset.entry_profile_metrics:
                ws.append([
                    row.document_id,
                    row.degree_name,
                    row.academic_year,
                    row.campus,
                    row.metric_name_std,
                    row.metric_name_raw,
                    row.metric_value,
                    row.unit,
                    row.source_pdf,
                ])

    def _write_subjects_sheet(self, ws, datasets):
        headers = [
            "document_id",
            "degree_name",
            "academic_year",
            "campus",
            "year_of_study",
            "subject_name",
            "subject_type_raw",
            "subject_type_std",
            "enrolled_total",
            "first_enrollment",
            "repeat_enrollment",
            "first_enrollment_performance",
            "passed",
            "failed",
            "not_presented",
            "performance_rate",
            "presentation_rate",
            "success_rate",
            "teaching_evaluation",
            "grade_ss",
            "grade_ap",
            "grade_nt",
            "grade_sb",
            "source_pdf",
        ]
        ws.append(headers)

        for dataset in datasets:
            for row in dataset.subject_results:
                ws.append([
                    row.document_id,
                    row.degree_name,
                    row.academic_year,
                    row.campus,
                    row.year_of_study,
                    row.subject_name,
                    row.subject_type_raw,
                    row.subject_type_std,
                    row.enrolled_total,
                    row.first_enrollment,
                    row.repeat_enrollment,
                    row.first_enrollment_performance,
                    row.passed,
                    row.failed,
                    row.not_presented,
                    row.performance_rate,
                    row.presentation_rate,
                    row.success_rate,
                    row.teaching_evaluation,
                    row.grade_ss,
                    row.grade_ap,
                    row.grade_nt,
                    row.grade_sb,
                    row.source_pdf,
                ])